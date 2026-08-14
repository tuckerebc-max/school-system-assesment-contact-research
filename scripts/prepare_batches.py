#!/usr/bin/env python3
"""Create stable district batches, a batch manifest, and a progress ledger."""

from __future__ import annotations

import argparse
import csv
import math
import sys
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_HEADERS = (
    "NCES District ID",
    "State District ID",
    "District Name",
    "County Name",
    "City",
    "State",
)

TIER1_LEDGER_COLUMNS = (
    "Superintendent / Chief Executive Status",
    "Chief Academic Officer / Teaching and Learning Status",
    "Assessment, Accountability, Research, and Evaluation Status",
    "Student / Scholar Services Status",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet", default="Districts results")
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--batch-size", type=int, default=25)
    parser.add_argument("--force", action="store_true")
    return parser.parse_args()


def clean_identifier(value: object, width: int | None = None) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()
    if width and text.isdigit():
        text = text.zfill(width)
    return text


def find_sheet(workbook, requested: str):
    if requested in workbook.sheetnames:
        return workbook[requested]
    matches = [name for name in workbook.sheetnames if name.casefold() == requested.casefold()]
    if len(matches) == 1:
        return workbook[matches[0]]
    raise ValueError(
        f"Worksheet {requested!r} was not found. Available: {', '.join(workbook.sheetnames)}"
    )


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    args = parse_args()
    if args.batch_size < 1:
        raise ValueError("--batch-size must be at least 1")
    if not args.workbook.exists():
        raise FileNotFoundError(args.workbook)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = args.output_dir / "batch-manifest.csv"
    ledger_path = args.output_dir / "district-ledger.csv"
    if not args.force and (manifest_path.exists() or ledger_path.exists()):
        raise FileExistsError(
            "Progress files already exist. Preserve them or rerun with --force after explicit reset approval."
        )

    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    sheet = find_sheet(workbook, args.sheet)
    header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in header_values]
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Missing required headers: {', '.join(missing)}")

    records: list[dict[str, object]] = []
    for worksheet_row, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not any(value not in (None, "") for value in values):
            continue
        row = {headers[index]: values[index] for index in range(len(headers))}
        row["NCES District ID"] = clean_identifier(row["NCES District ID"], 7)
        row["State District ID"] = clean_identifier(row["State District ID"])
        row["Source Row Number"] = worksheet_row
        row["District Ordinal"] = len(records) + 1
        row["Batch Number"] = math.ceil((len(records) + 1) / args.batch_size)
        records.append(row)

    if not records:
        raise ValueError("No district rows were found")
    nces_ids = [str(row["NCES District ID"]) for row in records]
    if "" in nces_ids:
        raise ValueError("Every district must have an NCES District ID")
    if len(nces_ids) != len(set(nces_ids)):
        raise ValueError("NCES District ID values are not unique")

    input_fields = ["Batch Number", "District Ordinal", "Source Row Number", *headers]
    batch_count = math.ceil(len(records) / args.batch_size)
    manifest_rows: list[dict[str, object]] = []
    for batch_number in range(1, batch_count + 1):
        batch_rows = [
            row for row in records if int(row["Batch Number"]) == batch_number
        ]
        batch_name = f"batch-{batch_number:03d}-input.csv"
        write_csv(args.output_dir / batch_name, input_fields, batch_rows)
        manifest_rows.append(
            {
                "Batch Number": batch_number,
                "Batch Size": len(batch_rows),
                "Start District Ordinal": batch_rows[0]["District Ordinal"],
                "End District Ordinal": batch_rows[-1]["District Ordinal"],
                "Start Source Row": batch_rows[0]["Source Row Number"],
                "End Source Row": batch_rows[-1]["Source Row Number"],
                "First NCES District ID": batch_rows[0]["NCES District ID"],
                "First District Name": batch_rows[0]["District Name"],
                "Last NCES District ID": batch_rows[-1]["NCES District ID"],
                "Last District Name": batch_rows[-1]["District Name"],
                "Input File": batch_name,
                "Status": "Queued",
                "Completed At": "",
                "Output Path": "",
            }
        )

    manifest_fields = list(manifest_rows[0].keys())
    write_csv(manifest_path, manifest_fields, manifest_rows)

    ledger_rows = []
    for row in records:
        ledger_rows.append(
            {
                "Batch Number": row["Batch Number"],
                "District Ordinal": row["District Ordinal"],
                "Source Row Number": row["Source Row Number"],
                "NCES District ID": row["NCES District ID"],
                "State District ID": row["State District ID"],
                "Input District Name": row["District Name"],
                "State": row["State"],
                "County Name": row["County Name"],
                "City": row["City"],
                "District Status": "Queued",
                **{column: "" for column in TIER1_LEDGER_COLUMNS},
                "Last Research Date": "",
                "QC Status": "Not Audited",
                "Output Path": "",
            }
        )
    ledger_fields = list(ledger_rows[0].keys())
    write_csv(ledger_path, ledger_fields, ledger_rows)

    print(
        f"Prepared {len(records)} districts in {batch_count} batches "
        f"({batch_count - 1} full batches plus a final batch of "
        f"{len(records) - (batch_count - 1) * args.batch_size})."
    )
    print(manifest_path)
    print(ledger_path)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
